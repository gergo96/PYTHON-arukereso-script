from bs4 import BeautifulSoup
import requests

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

######################################################

urlList = [] ### az URL címek || 25
numOfStores = [] ### egyes termékeket hányan árulják
nameOfProducts = []

priceList = []
storeList = []
######################################################

print("Írj be egy árukereső kategória URL címet")

urlname = input()
response = requests.get(urlname)
soup = BeautifulSoup(response.text, 'html.parser')
products = soup.find_all(class_="product-box clearfix")

###################################################### LINKEK MEGSZERZÉSE

for product in products:
    urlList.append(product.a["href"])
    print("URL hozzá adva a listához...")
    nameOfProducts.append(product.a["title"])
    print("Termék név hozzá adva a listához...")

###################################################### HÁNY BOLT ÁRULJA AZ EGYES TERMÉKEKET

for url in urlList:
    urlname = url
    response = requests.get(urlname)
    soup = BeautifulSoup(response.text, 'html.parser')

    prices = soup.find_all(class_="optoffer device-desktop")
    numOfStores.append(len(soup.find_all(class_="optoffer device-desktop")))
    for price in prices:
        try:
            priceList.append(price.find(class_='row-price').span.getText())
            print("Ár hozzá adva a listához...")
            storeList.append(price.div.div.img["alt"])
            print("Bolt név hozzá adva a listához...")
        except:
            storeList.append("Sajnos nem találtam...")
            print("valami nem volt jó, de azért listázom...")
            continue

lastRow = [numOfStores[0]]

#######################################################

for count in range(0,24):
    lastRow.append(lastRow[count]+numOfStores[count+1])


timeNow = datetime.datetime.now()
saveTime = str(timeNow.date()) + "_" + str(timeNow.strftime('%X')[0:5])
saveTime = saveTime.replace(":", "_")

wb = Workbook()

ws = wb.active
ws.title = "Árak"

ws["B1"] = "Bolt"
ws["C1"] = "Ár (Ft)"

nameFormat = Font(bold=True)

count_names = 0
for row in lastRow:
    ws.cell(row=row-numOfStores[count_names]+2, column=1).font = nameFormat
    ws.cell(row=row-numOfStores[count_names]+2, column=1).value = nameOfProducts[count_names]
    count_names+=1


counter = 0
for row in range(2,lastRow[len(lastRow)-1]+1):
    try:
        ws.cell(row=row, column=2).value = storeList[counter]
        ws.cell(row=row, column=3).value = int(priceList[counter].replace("Ft","").replace(" ",""))
        counter+=1
    except:
        print("valami nem volt jó, de azért listázom...")
        continue

wb.save(saveTime + "__Arukereso_Export.xlsx")