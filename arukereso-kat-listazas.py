from bs4 import BeautifulSoup
import requests

import datetime
from openpyxl import Workbook

timeNow = datetime.datetime.now()
saveTime = str(timeNow.date()) + "_" + str(timeNow.strftime('%X')[0:5])
saveTime = saveTime.replace(":", "_")

wb = Workbook()
ws = wb.active
ws.title = 'Kategória top25 Lista'

ws["B1"] = "Bolt"
ws["C1"] = "Ár"

lengthOfProducts = 0

lastRow = []

productNames = []
priceList = []
storeList = []
urlNames = []


print("Írj be egy árukereső kategória URL címet, vagy a visszalépéshez írd be, hogy 'back'")
urlname = input()

response = requests.get(urlname)

soup = BeautifulSoup(response.text, 'html.parser')

products = soup.find_all(class_="product-box clearfix")
#################
for each_product in range(0,len(products)):

    counter = 0
    for product in products:
        productNames.append(product.find(class_="name ulined-link").h2.a.getText())  # Termék név | 25db
        print(str(counter+1) + ". " + productNames[counter])
        counter+=1
        urlname = product.find(class_="name ulined-link").h2.a['href'] # Termék link |link
        urlNames.append(urlname)

    print(urlNames)

    for url in urlNames: #25 darab tehát 25ször jön be ide
        response = requests.get(url)

        soup = BeautifulSoup(response.text, 'html.parser')
        prices = soup.find_all(class_="optoffer device-desktop")
        row_counter = 1
        for price in prices:
            storeList.append(price.div.div.img['alt'])
            priceList.append(price.find(class_="row-price").span.getText())
            lastRow = len(prices)





#
# print("Írj be egy árukereső kategória URL címet, vagy a visszalépéshez írd be, hogy 'back'")
# urlname = input()
#
# response = requests.get(urlname)
#
# soup = BeautifulSoup(response.text, 'html.parser')
#
# counter = 0
# products = soup.find_all(class_="product-box clearfix")
# #################
# for product in products:
#     productNames.append(product.find(class_="name ulined-link").h2.a.getText())  # Termék név | 25db
#     print(str(counter+1) + ". " + productNames[counter])
#     counter+=1
#     urlname = product.find(class_="name ulined-link").h2.a['href'] # Termék link |link
#     urlNames.append()
#     print(urlname)
#     response = requests.get(urlname)
#     soup = BeautifulSoup(response.text, 'html.parser')
#
#     #Termék árak listázása
#     prices = soup.find_all(class_="optoffer device-desktop") #az összes ár+bolt
#     lastRow += len(prices)#hányan árulják
#
#
    # for row_count in range(0, len(productNames)):
    #     ws.cell(row=lastRow, column=1).value = productNames[row_count]
#
#
# wb.save(saveTime+"__Arukereso_Export.xlsx")
#
# #         #print(prices)
# # for price in prices:
#     storeList.append(price.div.div.img['alt'])
#     priceList.append(price.find(class_="row-price").span.getText())
#
#
#
# print("#######")
# print(lastRow)
#
#
#
#
#
#
#
# ########################################################################################################x
# #
# # timeNow = datetime.datetime.now()
# # saveTime = str(timeNow.date()) + "_" + str(timeNow.strftime('%X')[0:5])
# # saveTime = saveTime.replace(":", "_")
# #
# #
# # wb = Workbook()
# # ws = wb.active
# #
# # ws.title = 'Kategória top25 Lista'
# #
# # ws["B1"] = "Bolt"
# # ws["C1"] = "Ár"
# #
# # rowCounter = 3
# # for row_count in range(1,len(productNames)):
# #     ws.cell(row=row_count, column=1).value = productNames[row_count-1]
# #     for store in storeList:
# #         ws.cell(row=rowCounter, column=1).value = store
# #     for price in priceList:
# #         ws.cell(row=rowCounter,column=2).value = price
# #
# #
# # # wb.save(saveTime+"__Arukereso_Export.xlsx")

